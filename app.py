"""
Universal Crypto Wallet Balance Checker — Fast Edition
Run: python app.py  →  http://localhost:5000
pip install flask flask-cors requests pandas openpyxl
"""

import io, os, time
from concurrent.futures import ThreadPoolExecutor, as_completed
import requests, pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime, timezone
from flask import Flask, request, jsonify, send_from_directory, send_file
from flask_cors import CORS

app = Flask(__name__, static_folder="static")
CORS(app)

# ── shared HTTP session (connection pooling = faster) ─────────────────────────
SESSION = requests.Session()
SESSION.headers.update({"User-Agent": "CryptoChecker/2.0"})

# ── Coin registry ─────────────────────────────────────────────────────────────
COINS = {
    "BTC":  {"name":"Bitcoin",           "color":"#F7931A","dec":8,"cg":"bitcoin",              "type":"btc"},
    "LTC":  {"name":"Litecoin",          "color":"#BFBBBB","dec":8,"cg":"litecoin",             "type":"blockchair","chain":"litecoin"},
    "BCH":  {"name":"Bitcoin Cash",      "color":"#8DC351","dec":8,"cg":"bitcoin-cash",         "type":"blockchair","chain":"bitcoin-cash"},
    "BSV":  {"name":"Bitcoin SV",        "color":"#EAB300","dec":8,"cg":"bitcoin-sv",           "type":"blockchair","chain":"bitcoin-sv"},
    "DOGE": {"name":"Dogecoin",          "color":"#C2A633","dec":4,"cg":"dogecoin",             "type":"blockchair","chain":"dogecoin"},
    "ZEC":  {"name":"Zcash",             "color":"#F4B728","dec":8,"cg":"zcash",                "type":"blockchair","chain":"zcash"},
    "DASH": {"name":"Dash",              "color":"#008CE7","dec":8,"cg":"dash",                 "type":"blockchair","chain":"dash"},
    "DGB":  {"name":"DigiByte",          "color":"#0066CC","dec":8,"cg":"digibyte",             "type":"blockchair","chain":"digibyte"},
    "RVN":  {"name":"Ravencoin",         "color":"#384182","dec":8,"cg":"ravencoin",            "type":"blockchair","chain":"ravencoin"},
    "BTG":  {"name":"Bitcoin Gold",      "color":"#EEA109","dec":8,"cg":"bitcoin-gold",         "type":"blockchair","chain":"bitcoin-gold"},
    "ETH":  {"name":"Ethereum",          "color":"#627EEA","dec":6,"cg":"ethereum",             "type":"evm","rpc":"https://eth.llamarpc.com"},
    "ETC":  {"name":"Ethereum Classic",  "color":"#328332","dec":6,"cg":"ethereum-classic",     "type":"evm","rpc":"https://etc.etcdesktop.com"},
    "BNB":  {"name":"BNB (BSC)",         "color":"#F3BA2F","dec":6,"cg":"binancecoin",          "type":"evm","rpc":"https://bsc-dataseed.binance.org"},
    "MATIC":{"name":"Polygon",           "color":"#8247E5","dec":4,"cg":"matic-network",        "type":"evm","rpc":"https://polygon-rpc.com"},
    "AVAX": {"name":"Avalanche C-Chain", "color":"#E84142","dec":4,"cg":"avalanche-2",          "type":"evm","rpc":"https://api.avax.network/ext/bc/C/rpc"},
    "FTM":  {"name":"Fantom",            "color":"#1969FF","dec":4,"cg":"fantom",               "type":"evm","rpc":"https://rpc.ftm.tools"},
    "CRO":  {"name":"Cronos",            "color":"#002D74","dec":4,"cg":"crypto-com-chain",     "type":"evm","rpc":"https://evm.cronos.org"},
    "CELO": {"name":"Celo",              "color":"#35D07F","dec":4,"cg":"celo",                 "type":"evm","rpc":"https://forno.celo.org"},
    "ONE":  {"name":"Harmony",           "color":"#00AEE9","dec":4,"cg":"harmony",              "type":"evm","rpc":"https://api.harmony.one"},
    "XDAI": {"name":"Gnosis Chain",      "color":"#04795B","dec":4,"cg":"xdai",                 "type":"evm","rpc":"https://rpc.gnosischain.com"},
    "KLAY": {"name":"Klaytn",            "color":"#FF6B00","dec":4,"cg":"klay-token",           "type":"evm","rpc":"https://public-node-api.klaytnapi.com/v1/cypress"},
    "MOVR": {"name":"Moonriver",         "color":"#F2A007","dec":4,"cg":"moonriver",            "type":"evm","rpc":"https://rpc.api.moonriver.moonbeam.network"},
    "METIS":{"name":"Metis",             "color":"#00DACC","dec":4,"cg":"metis-token",          "type":"evm","rpc":"https://andromeda.metis.io/?owner=1088"},
    "FUSE": {"name":"Fuse",              "color":"#B5EC57","dec":4,"cg":"fuse-network-token",   "type":"evm","rpc":"https://rpc.fuse.io"},
    "IOTX": {"name":"IoTeX",             "color":"#00D4D5","dec":4,"cg":"iotex",                "type":"evm","rpc":"https://babel-api.mainnet.iotex.io"},
    "USDT": {"name":"Tether (ERC-20)",   "color":"#26A17B","dec":2,"cg":"tether",               "type":"erc20","contract":"0xdAC17F958D2ee523a2206206994597C13D831ec7","token_dec":6},
    "USDC": {"name":"USD Coin (ERC-20)", "color":"#2775CA","dec":2,"cg":"usd-coin",             "type":"erc20","contract":"0xA0b86991c6218b36c1d19D4a2e9Eb0cE3606eB48","token_dec":6},
    "DAI":  {"name":"Dai (ERC-20)",      "color":"#F5AC37","dec":2,"cg":"dai",                  "type":"erc20","contract":"0x6B175474E89094C44Da98b954EedeAC495271d0F","token_dec":18},
    "SHIB": {"name":"Shiba Inu (ERC-20)","color":"#FFA409","dec":0,"cg":"shiba-inu",            "type":"erc20","contract":"0x95aD61b0a150d79219dCF64E1E6Cc01f0B64C4cE","token_dec":18},
    "LINK": {"name":"Chainlink (ERC-20)","color":"#2A5ADA","dec":4,"cg":"chainlink",            "type":"erc20","contract":"0x514910771AF9Ca656af840dff83E8264EcF986CA","token_dec":18},
    "UNI":  {"name":"Uniswap (ERC-20)",  "color":"#FF007A","dec":4,"cg":"uniswap",              "type":"erc20","contract":"0x1f9840a85d5aF5bf1D1762F925BDADdC4201F984","token_dec":18},
    "WBTC": {"name":"Wrapped BTC",       "color":"#F7931A","dec":8,"cg":"wrapped-bitcoin",      "type":"erc20","contract":"0x2260FAC5E5542a773Aa44fBCfeDf7C193bc2C599","token_dec":8},
    "AAVE": {"name":"Aave (ERC-20)",     "color":"#B6509E","dec":4,"cg":"aave",                 "type":"erc20","contract":"0x7Fc66500c84A76Ad7e9c93437bFc5Ac33E2DDaE9","token_dec":18},
    "MKR":  {"name":"Maker (ERC-20)",    "color":"#1AAB9B","dec":4,"cg":"maker",                "type":"erc20","contract":"0x9f8F72aA9304c8B593d555F12eF6589cC3A579A2","token_dec":18},
    "GRT":  {"name":"The Graph",         "color":"#6747ED","dec":4,"cg":"the-graph",            "type":"erc20","contract":"0xc944E90C64B2c07662A292be6244BDf05Cda44a7","token_dec":18},
    "LDO":  {"name":"Lido DAO",          "color":"#F5A623","dec":4,"cg":"lido-dao",             "type":"erc20","contract":"0x5A98FcBEA516Cf06857215779Fd812CA3beF1B32","token_dec":18},
    "SOL":  {"name":"Solana",            "color":"#9945FF","dec":6,"cg":"solana",               "type":"solana"},
    "XRP":  {"name":"XRP",               "color":"#00AAE4","dec":6,"cg":"ripple",               "type":"xrp"},
    "ADA":  {"name":"Cardano",           "color":"#0033AD","dec":6,"cg":"cardano",              "type":"cardano"},
    "TRX":  {"name":"TRON",              "color":"#FF0013","dec":4,"cg":"tron",                 "type":"tron"},
    "XLM":  {"name":"Stellar Lumens",    "color":"#7D00FF","dec":4,"cg":"stellar",              "type":"stellar"},
    "XTZ":  {"name":"Tezos",             "color":"#2C7DF7","dec":4,"cg":"tezos",                "type":"tezos"},
    "ALGO": {"name":"Algorand",          "color":"#00B4D8","dec":4,"cg":"algorand",             "type":"algo"},
    "NEAR": {"name":"NEAR Protocol",     "color":"#00C08B","dec":4,"cg":"near",                 "type":"near"},
    "DOT":  {"name":"Polkadot",          "color":"#E6007A","dec":4,"cg":"polkadot",             "type":"subscan","network":"polkadot"},
    "KSM":  {"name":"Kusama",            "color":"#000000","dec":4,"cg":"kusama",               "type":"subscan","network":"kusama"},
    "ATOM": {"name":"Cosmos Hub",        "color":"#6F7390","dec":4,"cg":"cosmos",               "type":"cosmos","lcd":"https://rest.cosmos.directory/cosmoshub","denom":"uatom","div":1e6},
    "OSMO": {"name":"Osmosis",           "color":"#750BBB","dec":4,"cg":"osmosis",              "type":"cosmos","lcd":"https://rest.cosmos.directory/osmosis","denom":"uosmo","div":1e6},
    "JUNO": {"name":"Juno",              "color":"#F0827D","dec":4,"cg":"juno-network",         "type":"cosmos","lcd":"https://rest.cosmos.directory/juno","denom":"ujuno","div":1e6},
    "KAVA": {"name":"Kava",              "color":"#FF564F","dec":4,"cg":"kava",                 "type":"cosmos","lcd":"https://rest.cosmos.directory/kava","denom":"ukava","div":1e6},
    "SCRT": {"name":"Secret Network",    "color":"#1B1B1B","dec":4,"cg":"secret",               "type":"cosmos","lcd":"https://rest.cosmos.directory/secretnetwork","denom":"uscrt","div":1e6},
    "VET":  {"name":"VeChain",           "color":"#15BDFF","dec":2,"cg":"vechain",              "type":"vet"},
    "HBAR": {"name":"Hedera",            "color":"#222222","dec":4,"cg":"hedera-hashgraph",     "type":"hedera"},
    "EOS":  {"name":"EOS",               "color":"#443F54","dec":4,"cg":"eos",                  "type":"eos"},
    "WAVES":{"name":"Waves",             "color":"#0055FF","dec":4,"cg":"waves",                "type":"waves"},
    "XNO":  {"name":"Nano",              "color":"#4A90E2","dec":4,"cg":"nano",                 "type":"nano"},
    "XMR":  {"name":"Monero (Private)",  "color":"#FF6600","dec":6,"cg":"monero",               "type":"private"},
    "ICP":  {"name":"Internet Computer", "color":"#292A2E","dec":4,"cg":"internet-computer",    "type":"icp"},
    "FIL":  {"name":"Filecoin",          "color":"#42C1CA","dec":4,"cg":"filecoin",             "type":"filecoin"},
    "APT":  {"name":"Aptos",             "color":"#2DD8A3","dec":4,"cg":"aptos",                "type":"aptos"},
    "SUI":  {"name":"Sui",               "color":"#6FBCF0","dec":4,"cg":"sui",                  "type":"sui"},
    "TON":  {"name":"Toncoin",           "color":"#0098EA","dec":4,"cg":"the-open-network",     "type":"ton"},
    "KAS":  {"name":"Kaspa",             "color":"#49EACB","dec":4,"cg":"kaspa",                "type":"kaspa"},
    "EGLD": {"name":"MultiversX",        "color":"#1D9BF0","dec":4,"cg":"elrond-erd-2",         "type":"multiversx"},
    "FLOW": {"name":"Flow",              "color":"#00EF8B","dec":4,"cg":"flow",                 "type":"flow"},
}

# ── Price cache ───────────────────────────────────────────────────────────────
_pcache = {"data": {}, "ts": 0.0}

def get_prices(symbols):
    now = time.time()
    needed = [s for s in symbols if s not in _pcache["data"] or now - _pcache["ts"] > 60]
    if needed:
        cg_ids = ",".join(COINS[s]["cg"] for s in needed if s in COINS and "cg" in COINS[s])
        if cg_ids:
            try:
                r = SESSION.get("https://api.coingecko.com/api/v3/simple/price",
                    params={"ids": cg_ids, "vs_currencies": "usd"}, timeout=10)
                r.raise_for_status()
                id2sym = {COINS[s]["cg"]: s for s in COINS}
                for cg_id, vals in r.json().items():
                    sym = id2sym.get(cg_id)
                    if sym:
                        _pcache["data"][sym] = vals["usd"]
                _pcache["ts"] = now
            except Exception as e:
                print(f"[price] {e}")
    return {s: _pcache["data"].get(s, 0.0) for s in symbols}

# ── Balance fetchers (all use shared SESSION) ─────────────────────────────────

def _get(url, **kw):
    kw.setdefault("timeout", 8)
    return SESSION.get(url, **kw)

def _post(url, **kw):
    kw.setdefault("timeout", 8)
    return SESSION.post(url, **kw)

def _btc(addr):
    for url in [f"https://blockstream.info/api/address/{addr}",
                f"https://mempool.space/api/address/{addr}"]:
        try:
            r = _get(url)
            if r.status_code == 200:
                d = r.json()
                return (d["chain_stats"]["funded_txo_sum"] - d["chain_stats"]["spent_txo_sum"]) / 1e8
        except Exception:
            pass
    try:
        r = _get(f"https://api.blockchair.com/bitcoin/dashboards/address/{addr}")
        if r.status_code == 200:
            d = r.json()["data"]
            if d: return list(d.values())[0]["address"]["balance"] / 1e8
    except Exception:
        pass
    return None

def _blockchair(addr, chain):
    try:
        r = _get(f"https://api.blockchair.com/{chain}/dashboards/address/{addr}")
        if r.status_code == 200:
            d = r.json()["data"]
            if d: return list(d.values())[0]["address"]["balance"] / 1e8
    except Exception:
        pass
    return None

def _evm(addr, rpc):
    try:
        r = _post(rpc, json={"jsonrpc":"2.0","method":"eth_getBalance","params":[addr,"latest"],"id":1})
        if r.status_code == 200:
            return int(r.json().get("result","0x0"), 16) / 1e18
    except Exception:
        pass
    return None

def _erc20(addr, contract, token_dec):
    padded = addr.lower().replace("0x","").zfill(64)
    data   = "0x70a08231" + padded
    try:
        r = _post("https://eth.llamarpc.com",
            json={"jsonrpc":"2.0","method":"eth_call","params":[{"to":contract,"data":data},"latest"],"id":1})
        if r.status_code == 200:
            result = r.json().get("result","0x0")
            if result and result != "0x":
                return int(result, 16) / (10**token_dec)
    except Exception:
        pass
    return None

def _solana(addr):
    for rpc in ["https://api.mainnet-beta.solana.com","https://solana-api.projectserum.com"]:
        try:
            r = _post(rpc, json={"jsonrpc":"2.0","id":1,"method":"getBalance","params":[addr]})
            if r.status_code == 200:
                v = r.json().get("result",{}).get("value")
                if v is not None: return v / 1e9
        except Exception:
            pass
    return None

def _xrp(addr):
    try:
        r = _post("https://xrplcluster.com",
            json={"method":"account_info","params":[{"account":addr,"ledger_index":"current"}]})
        if r.status_code == 200:
            drops = r.json().get("result",{}).get("account_data",{}).get("Balance")
            if drops: return int(drops) / 1e6
    except Exception:
        pass
    return None

def _cardano(addr):
    try:
        r = _post("https://api.koios.rest/api/v1/address_info", json={"_addresses":[addr]})
        if r.status_code == 200:
            data = r.json()
            if data: return int(data[0].get("balance",0)) / 1e6
    except Exception:
        pass
    return None

def _tron(addr):
    try:
        r = _post("https://api.trongrid.io/wallet/getaccount", json={"address":addr,"visible":True})
        if r.status_code == 200:
            return r.json().get("balance",0) / 1e6
    except Exception:
        pass
    return None

def _stellar(addr):
    try:
        r = _get(f"https://horizon.stellar.org/accounts/{addr}")
        if r.status_code == 200:
            for b in r.json().get("balances",[]):
                if b.get("asset_type") == "native":
                    return float(b["balance"])
    except Exception:
        pass
    return None

def _tezos(addr):
    try:
        r = _get(f"https://api.tzkt.io/v1/accounts/{addr}/balance")
        if r.status_code == 200: return r.json() / 1e6
    except Exception:
        pass
    return None

def _algo(addr):
    try:
        r = _get(f"https://mainnet-api.algonode.cloud/v2/accounts/{addr}")
        if r.status_code == 200: return r.json().get("amount",0) / 1e6
    except Exception:
        pass
    return None

def _near(addr):
    try:
        r = _post("https://rpc.mainnet.near.org",
            json={"jsonrpc":"2.0","id":"x","method":"query",
                  "params":{"request_type":"view_account","finality":"final","account_id":addr}})
        if r.status_code == 200:
            yocto = r.json().get("result",{}).get("amount","0")
            return int(yocto) / 1e24
    except Exception:
        pass
    return None

def _subscan(addr, network):
    try:
        r = _post(f"https://{network}.api.subscan.io/api/open/account",
            json={"address":addr}, headers={"Content-Type":"application/json"})
        if r.status_code == 200:
            bal = r.json().get("data",{}).get("balance")
            if bal is not None: return float(bal)
    except Exception:
        pass
    return None

def _cosmos(addr, lcd, denom, div):
    try:
        r = _get(f"{lcd}/cosmos/bank/v1beta1/balances/{addr}")
        if r.status_code == 200:
            for b in r.json().get("balances",[]):
                if b.get("denom") == denom:
                    return int(b["amount"]) / div
            return 0.0
    except Exception:
        pass
    return None

def _vet(addr):
    for url in [f"https://mainnet.veblocks.net/accounts/{addr}",
                f"https://vethor-node.vechain.com/accounts/{addr}"]:
        try:
            r = _get(url)
            if r.status_code == 200:
                return int(r.json().get("balance","0x0"), 16) / 1e18
        except Exception:
            pass
    return None

def _hedera(addr):
    try:
        acc = addr if "." in addr else addr
        r = _get(f"https://mainnet-public.mirrornode.hedera.com/api/v1/accounts/{acc}")
        if r.status_code == 200:
            return r.json().get("balance",{}).get("balance",0) / 1e8
    except Exception:
        pass
    return None

def _eos(addr):
    try:
        r = _post("https://eos.greymass.com/v1/chain/get_account", json={"account_name":addr})
        if r.status_code == 200:
            bal = r.json().get("core_liquid_balance","0 EOS")
            return float(bal.split()[0])
    except Exception:
        pass
    return None

def _waves(addr):
    try:
        r = _get(f"https://nodes.wavesnodes.com/addresses/balance/{addr}")
        if r.status_code == 200: return r.json().get("balance",0) / 1e8
    except Exception:
        pass
    return None

def _nano(addr):
    try:
        r = _post("https://mynano.ninja/api/node",
            json={"action":"account_balance","account":addr})
        if r.status_code == 200:
            return int(r.json().get("balance","0")) / (10**30)
    except Exception:
        pass
    return None

def _icp(addr):
    try:
        r = _get(f"https://ic-api.internetcomputer.org/api/v3/accounts/{addr}")
        if r.status_code == 200:
            return r.json().get("balance",0) / 1e8
    except Exception:
        pass
    return None

def _filecoin(addr):
    try:
        r = _post("https://api.node.glif.io/rpc/v0",
            json={"jsonrpc":"2.0","method":"Filecoin.WalletBalance","params":[addr],"id":1})
        if r.status_code == 200:
            return int(r.json().get("result","0")) / 1e18
    except Exception:
        pass
    return None

def _aptos(addr):
    try:
        r = _get(f"https://fullnode.mainnet.aptoslabs.com/v1/accounts/{addr}/resource/0x1::coin::CoinStore<0x1::aptos_coin::AptosCoin>")
        if r.status_code == 200:
            return int(r.json()["data"]["coin"]["value"]) / 1e8
    except Exception:
        pass
    return None

def _sui(addr):
    try:
        r = _post("https://fullnode.mainnet.sui.io:443",
            json={"jsonrpc":"2.0","id":1,"method":"suix_getBalance","params":[addr,"0x2::sui::SUI"]})
        if r.status_code == 200:
            return int(r.json().get("result",{}).get("totalBalance","0")) / 1e9
    except Exception:
        pass
    return None

def _ton(addr):
    try:
        r = _get(f"https://tonapi.io/v2/accounts/{addr}")
        if r.status_code == 200:
            return r.json().get("balance",0) / 1e9
    except Exception:
        pass
    return None

def _kaspa(addr):
    try:
        r = _get(f"https://api.kaspa.org/addresses/{addr}/balance")
        if r.status_code == 200:
            return r.json().get("balance",0) / 1e8
    except Exception:
        pass
    return None

def _multiversx(addr):
    try:
        r = _get(f"https://api.multiversx.com/accounts/{addr}")
        if r.status_code == 200:
            return int(r.json().get("balance","0")) / 1e18
    except Exception:
        pass
    return None

def _flow(addr):
    try:
        r = _get(f"https://rest-mainnet.onflow.org/v1/accounts/{addr}")
        if r.status_code == 200:
            return int(r.json().get("balance","0")) / 1e8
    except Exception:
        pass
    return None

# ── Dispatcher ────────────────────────────────────────────────────────────────
def get_balance(address, coin):
    coin = coin.upper().strip()
    meta = COINS.get(coin)
    if not meta: return None
    t = meta["type"]
    try:
        if t == "btc":         return _btc(address)
        if t == "blockchair":  return _blockchair(address, meta["chain"])
        if t == "evm":         return _evm(address, meta["rpc"])
        if t == "erc20":       return _erc20(address, meta["contract"], meta["token_dec"])
        if t == "solana":      return _solana(address)
        if t == "xrp":         return _xrp(address)
        if t == "cardano":     return _cardano(address)
        if t == "tron":        return _tron(address)
        if t == "stellar":     return _stellar(address)
        if t == "tezos":       return _tezos(address)
        if t == "algo":        return _algo(address)
        if t == "near":        return _near(address)
        if t == "subscan":     return _subscan(address, meta["network"])
        if t == "cosmos":      return _cosmos(address, meta["lcd"], meta["denom"], meta["div"])
        if t == "vet":         return _vet(address)
        if t == "hedera":      return _hedera(address)
        if t == "eos":         return _eos(address)
        if t == "waves":       return _waves(address)
        if t == "nano":        return _nano(address)
        if t == "icp":         return _icp(address)
        if t == "filecoin":    return _filecoin(address)
        if t == "aptos":       return _aptos(address)
        if t == "sui":         return _sui(address)
        if t == "ton":         return _ton(address)
        if t == "kaspa":       return _kaspa(address)
        if t == "multiversx":  return _multiversx(address)
        if t == "flow":        return _flow(address)
        if t == "private":     return None
    except Exception as e:
        print(f"[balance] {coin} {e}")
    return None

# ── Helpers ───────────────────────────────────────────────────────────────────
def find_addr_col(df, hint=""):
    if hint and hint in df.columns: return hint
    for c in df.columns:
        if any(k in c.lower() for k in ("address","wallet","addr")): return c
    return df.columns[0]

def _process(addr, coin, prices):
    meta  = COINS.get(coin, {})
    price = prices.get(coin, 0)
    if not addr or addr.lower() in ("nan","none",""):
        return {"address":addr,"coin":coin,"balance":None,"usd":None,"error":"empty"}
    if coin not in COINS:
        return {"address":addr,"coin":coin,"balance":None,"usd":None,"error":"unsupported"}
    if COINS[coin]["type"] == "private":
        return {"address":addr,"coin":coin,"balance":None,"usd":None,"error":"private"}
    bal = get_balance(addr, coin)
    if bal is None:
        return {"address":addr,"coin":coin,"balance":None,"usd":None,"error":"fetch_failed"}
    return {"address":addr,"coin":coin,"name":meta.get("name",coin),
            "color":meta.get("color","#888"),
            "balance":round(bal, meta.get("dec",6)),
            "usd":round(bal*price, 2), "price":price, "error":None}

def write_excel(filepath, df, addr_col):
    buf = io.BytesIO()
    df.to_excel(buf, index=False, engine="openpyxl")
    buf.seek(0)
    wb = load_workbook(buf)
    ws = wb.active
    H_FILL = PatternFill("solid", start_color="0D1117")
    H_FONT = Font(bold=True, color="F0C040", name="Consolas", size=10)
    ALT    = PatternFill("solid", start_color="F0F4FF")
    NORM   = Font(name="Consolas", size=10)
    ERR    = Font(name="Consolas", size=10, color="C00000")
    for cell in ws[1]:
        cell.fill = H_FILL; cell.font = H_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for ri in range(2, len(df)+2):
        for cell in ws[ri]:
            cell.font = NORM
            if ri % 2 == 0: cell.fill = ALT
            cell.alignment = Alignment(
                horizontal="right" if cell.column > 2 else "left", vertical="center")
            if str(cell.value) in ("ERROR","UNSUPPORTED","PRIVATE"): cell.font = ERR
    for ci in range(1, ws.max_column+1):
        ws.column_dimensions[get_column_letter(ci)].width = 20
    try:
        ai = list(df.columns).index(addr_col) + 1
        ws.column_dimensions[get_column_letter(ai)].width = 48
    except Exception:
        pass
    ws.freeze_panes = "A2"
    wb.save(filepath)
    return buf   # also return bytes for in-memory download

def build_excel_bytes(df, addr_col):
    """Return styled xlsx as BytesIO (for download endpoint)."""
    buf = io.BytesIO()
    df.to_excel(buf, index=False, engine="openpyxl")
    buf.seek(0)
    wb = load_workbook(buf)
    ws = wb.active
    H_FILL = PatternFill("solid", start_color="0D1117")
    H_FONT = Font(bold=True, color="F0C040", name="Consolas", size=10)
    ALT    = PatternFill("solid", start_color="F0F4FF")
    NORM   = Font(name="Consolas", size=10)
    ERR    = Font(name="Consolas", size=10, color="C00000")
    for cell in ws[1]:
        cell.fill = H_FILL; cell.font = H_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for ri in range(2, len(df)+2):
        for cell in ws[ri]:
            cell.font = NORM
            if ri % 2 == 0: cell.fill = ALT
            cell.alignment = Alignment(
                horizontal="right" if cell.column > 2 else "left", vertical="center")
            if str(cell.value) in ("ERROR","UNSUPPORTED","PRIVATE"): cell.font = ERR
    for ci in range(1, ws.max_column+1):
        ws.column_dimensions[get_column_letter(ci)].width = 20
    try:
        ai = list(df.columns).index(addr_col) + 1
        ws.column_dimensions[get_column_letter(ai)].width = 48
    except Exception:
        pass
    ws.freeze_panes = "A2"
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out

# ── In-memory last excel result (for download) ────────────────────────────────
_last_excel = {"df": None, "addr_col": None, "filename": "results.xlsx"}

# ── Flask routes ──────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return send_from_directory("static", "index.html")

@app.route("/api/coins")
def api_coins():
    return jsonify([
        {"symbol":s, "name":m["name"], "color":m["color"],
         "supported": m["type"] != "private"}
        for s, m in COINS.items()
    ])

@app.route("/api/prices")
def api_prices():
    return jsonify(get_prices(list(COINS.keys())))

@app.route("/api/price/<symbol>")
def api_price_single(symbol):
    """Fast single-coin price lookup."""
    sym = symbol.upper()
    prices = get_prices([sym])
    return jsonify({"symbol": sym, "price": prices.get(sym, 0)})

@app.route("/api/single", methods=["POST"])
def api_single():
    d    = request.json or {}
    addr = (d.get("address") or "").strip()
    coin = (d.get("coin") or "BTC").upper()
    if not addr: return jsonify({"error": "No address provided"}), 400
    if coin not in COINS: return jsonify({"error": f"Unknown coin: {coin}"}), 400
    if COINS[coin]["type"] == "private":
        return jsonify({"error": f"{coin} is privacy-focused — balance not publicly queryable"}), 400
    prices = get_prices([coin])
    result = _process(addr, coin, prices)
    if result["error"]:
        return jsonify({"error": f"Could not fetch {coin} balance — check address/network"}), 502
    result["checked"] = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    return jsonify(result)

@app.route("/api/bulk", methods=["POST"])
def api_bulk():
    """Parallel processing — up to 10x faster than sequential."""
    d       = request.json or {}
    entries = d.get("entries", [])
    if not entries: return jsonify({"error": "No entries"}), 400

    syms   = list({e.get("coin","BTC").upper() for e in entries})
    prices = get_prices(syms)
    now    = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

    # Run all fetches in parallel (max 12 threads)
    results = [None] * len(entries)
    with ThreadPoolExecutor(max_workers=min(12, len(entries))) as ex:
        futures = {
            ex.submit(_process,
                      str(e.get("address","")).strip(),
                      str(e.get("coin","BTC")).upper(),
                      prices): i
            for i, e in enumerate(entries)
        }
        for fut in as_completed(futures):
            idx = futures[fut]
            try:
                results[idx] = fut.result()
            except Exception as exc:
                e = entries[idx]
                results[idx] = {"address": e.get("address",""), "coin": e.get("coin",""),
                                "balance": None, "usd": None, "error": str(exc)}

    return jsonify({"prices": prices, "checked": now, "results": results})

@app.route("/api/excel", methods=["POST"])
def api_excel():
    d        = request.json or {}
    filepath = (d.get("path") or "").strip()
    col_hint = (d.get("column") or "").strip()
    coin_col = (d.get("coin_column") or "").strip()
    def_coin = (d.get("default_coin") or "BTC").upper()

    if not filepath: return jsonify({"error": "No file path"}), 400
    if not os.path.isfile(filepath): return jsonify({"error": f"File not found: {filepath}"}), 404
    try:
        df = pd.read_excel(filepath, dtype=str)
    except Exception as e:
        return jsonify({"error": str(e)}), 400
    if df.empty: return jsonify({"error": "File is empty"}), 400

    addr_col   = find_addr_col(df, col_hint)
    now_str    = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    coins_list = [
        str(row.get(coin_col, def_coin) if coin_col and coin_col in df.columns else def_coin
            ).upper().strip()
        for _, row in df.iterrows()
    ]
    prices = get_prices(list(set(coins_list)))

    # Build task list
    tasks = []
    for i, (_, row) in enumerate(df.iterrows()):
        addr = str(row.get(addr_col, "")).strip()
        coin = coins_list[i]
        tasks.append((i, addr, coin))

    row_results = [None] * len(tasks)

    # Parallel fetch
    with ThreadPoolExecutor(max_workers=min(12, len(tasks))) as ex:
        futures = {ex.submit(_process, addr, coin, prices): i
                   for i, addr, coin in tasks}
        for fut in as_completed(futures):
            idx = futures[fut]
            try:
                row_results[idx] = fut.result()
            except Exception as exc:
                _, addr, coin = tasks[idx]
                row_results[idx] = {"address": addr, "coin": coin,
                                    "balance": None, "usd": None, "error": str(exc)}

    r_coin, r_bal, r_usd, r_ts = [], [], [], []
    for res in row_results:
        coin = res["coin"]
        r_coin.append(coin); r_ts.append(now_str)
        if res["error"]:
            label = {"empty":"N/A","private":"PRIVATE","unsupported":"UNSUPPORTED"}.get(res["error"],"ERROR")
            r_bal.append(label); r_usd.append(label)
        else:
            r_bal.append(res["balance"]); r_usd.append(res["usd"])

    df["Coin"]             = r_coin
    df["Balance"]          = r_bal
    df["USD Value"]        = r_usd
    df["Checked At (UTC)"] = r_ts

    # Save to disk
    write_excel(filepath, df, addr_col)

    # Keep in memory for download
    _last_excel["df"]       = df.copy()
    _last_excel["addr_col"] = addr_col
    fn = os.path.basename(filepath).replace(".xlsx","") + "_results.xlsx"
    _last_excel["filename"] = fn

    return jsonify({
        "prices": prices, "checked": now_str,
        "file": filepath, "column": addr_col,
        "results": row_results,
        "download_ready": True,
    })

@app.route("/api/excel/download")
def api_excel_download():
    """Stream the last processed Excel file as a download."""
    if _last_excel["df"] is None:
        return jsonify({"error": "No results yet — run a check first"}), 404
    buf = build_excel_bytes(_last_excel["df"], _last_excel["addr_col"])
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=_last_excel["filename"],
    )

if __name__ == "__main__":
    print("\n  ◆  Universal Crypto Wallet Balance Checker  (Fast Edition)")
    print("  ─────────────────────────────────────────────────────────")
    print(f"  {len(COINS)} cryptocurrencies | parallel fetching | live prices")
    print("  Open in browser →  http://localhost:5000\n")
    app.run(debug=False, port=5000, threaded=True)
